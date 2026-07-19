import asyncio
import base64
import io
import json
import logging
import zipfile
from collections.abc import AsyncGenerator

from fastapi import HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from src.api.schemas.request import FewShotExample, GradingRequest, InlineGradingRequest
from src.api.schemas.response import GradingResponse
from src.services.grading_service import GradingService

logger = logging.getLogger(__name__)


def _serialize_examples(
    examples: list[FewShotExample] | None,
) -> list[dict] | None:
    if not examples:
        return None
    return [ex.model_dump() for ex in examples]


class GradingController:
    """Handles HTTP-level concerns for grading endpoints.

    Responsible for decoding uploads, dispatching to GradingService,
    and formatting responses (JSON or Excel).
    """

    def __init__(self, service: GradingService, batch_concurrency: int = 5) -> None:
        self.service = service
        self.batch_concurrency = max(1, batch_concurrency)

    async def grade_inline(self, request: InlineGradingRequest) -> GradingResponse:
        logger.info("grade_inline — with_reason=%s", request.with_reason)

        internal = GradingRequest(
            problem_description=request.problems,
            student_code=request.code,
            rubric=request.rubric,
            with_reason=request.with_reason,
            few_shot_examples=_serialize_examples(request.few_shot_examples),
        )
        result = await self.service.grade(internal)

        if not request.with_reason:
            result.reasoning = None

        return result

    async def grade_file(
        self,
        problems: str,
        code: UploadFile,
        rubric: str | None,
        with_reason: bool,
        few_shot_examples: list[dict] | None = None,
    ) -> GradingResponse:
        logger.info(
            "grade_file — filename=%s, with_reason=%s", code.filename, with_reason
        )

        try:
            content = (await code.read()).decode("utf-8")
        except UnicodeDecodeError:
            raise HTTPException(
                status_code=400, detail="File could not be decoded as UTF-8 text"
            )

        internal = GradingRequest(
            problem_description=problems,
            student_code=content,
            rubric=rubric,
            with_reason=with_reason,
            few_shot_examples=few_shot_examples,
        )
        result = await self.service.grade(internal)

        if not with_reason:
            result.reasoning = None

        return result
    
    async def grade_batch(
        self,
        problems: str,
        files: UploadFile,
        rubric: str | None,
        with_reason: bool,
        few_shot_examples: list[dict] | None = None,
    ) -> StreamingResponse:
        logger.info(
            "grade_batch — zipfile=%s, with_reason=%s", files.filename, with_reason
        )

        zip_bytes = await files.read()
        try:
            zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
        except zipfile.BadZipFile:
            raise HTTPException(
                status_code=400, detail="Uploaded file is not a valid zip archive"
            )

        code_files = [
            name for name in zf.namelist() if self._is_supported_zip_entry(name)
        ]
        if not code_files:
            raise HTTPException(status_code=400, detail="Zip archive contains no files")

        logger.info("Batch contains %d file(s)", len(code_files))

        results: list[dict | None] = [None] * len(code_files)
        pending: list[tuple[int, str, str]] = []
        for idx, filename in enumerate(code_files):
            try:
                content = zf.read(filename).decode("utf-8")
            except UnicodeDecodeError:
                results[idx] = {
                    "filename": filename,
                    "error": f"Could not decode {filename} as UTF-8",
                }
                continue

            pending.append((idx, filename, content))

        semaphore = asyncio.Semaphore(self.batch_concurrency)
        tasks = [
            asyncio.create_task(
                self._grade_single_submission(
                    idx=idx,
                    filename=filename,
                    content=content,
                    problems=problems,
                    rubric=rubric,
                    with_reason=with_reason,
                    semaphore=semaphore,
                    few_shot_examples=few_shot_examples,
                )
            )
            for idx, filename, content in pending
        ]

        for task in asyncio.as_completed(tasks):
            idx, entry = await task
            results[idx] = entry

        excel_bytes = self._build_excel(
            [r for r in results if r is not None], with_reason=with_reason
        )

        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=grading_results.xlsx"
            },
        )

    async def grade_inline_stream(
        self, request: InlineGradingRequest
    ) -> AsyncGenerator[dict, None]:
        """SSE generator for inline grading — emits one progress event per pipeline step,
        then a final 'result' event with the full GradingResponse."""

        logger.info("grade_inline_stream — with_reason=%s", request.with_reason)

        internal = GradingRequest(
            problem_description=request.problems,
            student_code=request.code,
            rubric=request.rubric,
            with_reason=request.with_reason,
            few_shot_examples=_serialize_examples(request.few_shot_examples),
        )

        queue: asyncio.Queue[dict | None] = asyncio.Queue()

        async def on_progress_queued(step: int, total: int, message: str) -> None:
            await queue.put(
                {
                    "data": json.dumps(
                        {
                            "type": "progress",
                            "step": step,
                            "total": total,
                            "message": message,
                        }
                    )
                }
            )

        async def run_pipeline() -> None:
            try:
                result = await self.service.grade(
                    internal, on_progress=on_progress_queued
                )

                if not request.with_reason:
                    result.reasoning = None

                await queue.put(
                    {
                        "data": json.dumps(
                            {"type": "result", "data": result.model_dump()}
                        )
                    }
                )
            except Exception as exc:
                logger.error("grade_inline_stream error: %s", exc)

                await queue.put(
                    {"data": json.dumps({"type": "error", "message": str(exc)})}
                )
            finally:
                await queue.put(None)  # sentinel

        task = asyncio.create_task(run_pipeline())
        while True:
            item = await queue.get()
            if item is None:
                break
            yield item

        await task

    async def grade_file_stream(
        self,
        problems: str,
        code: UploadFile,
        rubric: str | None,
        with_reason: bool,
        few_shot_examples: list[dict] | None = None,
    ) -> AsyncGenerator[dict, None]:
        """SSE generator for file grading — same event shape as grade_inline_stream."""

        logger.info(
            "grade_file_stream — filename=%s, with_reason=%s",
            code.filename,
            with_reason,
        )

        try:
            content = (await code.read()).decode("utf-8")
        except UnicodeDecodeError:
            yield {
                "data": json.dumps(
                    {
                        "type": "error",
                        "message": "File could not be decoded as UTF-8 text",
                    }
                )
            }
            return

        internal = GradingRequest(
            problem_description=problems,
            student_code=content,
            rubric=rubric,
            with_reason=with_reason,
            few_shot_examples=few_shot_examples,
        )

        queue: asyncio.Queue[dict | None] = asyncio.Queue()

        async def on_progress(step: int, total: int, message: str) -> None:
            await queue.put(
                {
                    "data": json.dumps(
                        {
                            "type": "progress",
                            "step": step,
                            "total": total,
                            "message": message,
                        }
                    )
                }
            )

        async def run_pipeline() -> None:
            try:
                result = await self.service.grade(internal, on_progress=on_progress)

                if not with_reason:
                    result.reasoning = None

                await queue.put(
                    {
                        "data": json.dumps(
                            {"type": "result", "data": result.model_dump()}
                        )
                    }
                )
            except Exception as exc:
                logger.error("grade_file_stream error: %s", exc)

                await queue.put(
                    {"data": json.dumps({"type": "error", "message": str(exc)})}
                )
            finally:
                await queue.put(None)

        task = asyncio.create_task(run_pipeline())
        while True:
            item = await queue.get()
            if item is None:
                break
            yield item

        await task

    async def grade_batch_stream(
        self,
        problems: str,
        files: UploadFile,
        rubric: str | None,
        with_reason: bool,
        few_shot_examples: list[dict] | None = None,
    ) -> AsyncGenerator[dict, None]:
        """SSE event generator: emits progress per file, then a final 'complete' event
        containing the Excel as base64 so the client can trigger the download."""

        logger.info(
            "grade_batch_stream — zipfile=%s, with_reason=%s",
            files.filename,
            with_reason,
        )

        zip_bytes = await files.read()
        try:
            zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
        except zipfile.BadZipFile:
            yield {
                "data": json.dumps(
                    {
                        "type": "error",
                        "message": "Uploaded file is not a valid zip archive",
                    }
                )
            }
            return

        code_files = [
            name for name in zf.namelist() if self._is_supported_zip_entry(name)
        ]
        if not code_files:
            yield {
                "data": json.dumps(
                    {"type": "error", "message": "Zip archive contains no files"}
                )
            }
            return

        total = len(code_files)
        logger.info("Batch stream: %d file(s)", total)

        results: list[dict | None] = [None] * total
        queue: asyncio.Queue[tuple[int, dict]] = asyncio.Queue()
        semaphore = asyncio.Semaphore(self.batch_concurrency)

        async def grade_and_notify(idx: int, filename: str, content: str) -> None:
            _, entry = await self._grade_single_submission(
                idx=idx,
                filename=filename,
                content=content,
                problems=problems,
                rubric=rubric,
                with_reason=with_reason,
                semaphore=semaphore,
                few_shot_examples=few_shot_examples,
            )
            results[idx] = entry
            await queue.put((idx, entry))

        pending = []
        for idx, filename in enumerate(code_files):
            try:
                content = zf.read(filename).decode("utf-8")
            except UnicodeDecodeError:
                entry = {
                    "filename": filename,
                    "error": f"Could not decode {filename} as UTF-8",
                }
                results[idx] = entry

                await queue.put((idx, entry))
                continue

            pending.append(
                asyncio.create_task(grade_and_notify(idx, filename, content))
            )

        for done_count in range(1, total + 1):
            _, entry = await queue.get()
            event = {
                "type": "progress",
                "done": done_count,
                "total": total,
                "filename": entry["filename"],
                "score": entry.get("score"),
                "error": entry.get("error"),
            }

            logger.debug(
                "Batch progress %d/%d — %s", done_count, total, entry["filename"]
            )

            yield {"data": json.dumps(event)}

        if pending:
            await asyncio.gather(*pending)

        excel_bytes = self._build_excel(
            [r for r in results if r is not None], with_reason=with_reason
        )
        excel_b64 = base64.b64encode(excel_bytes).decode()

        yield {"data": json.dumps({"type": "complete", "excel": excel_b64})}

        logger.info("Batch stream complete — %d file(s) graded", total)

    async def _grade_single_submission(
        self,
        idx: int,
        filename: str,
        content: str,
        problems: str,
        rubric: str | None,
        with_reason: bool,
        semaphore: asyncio.Semaphore,
        few_shot_examples: list[dict] | None = None,
    ) -> tuple[int, dict]:
        internal = GradingRequest(
            problem_description=problems,
            student_code=content,
            rubric=rubric,
            with_reason=with_reason,
            few_shot_examples=few_shot_examples,
        )

        async with semaphore:
            try:
                logger.info("Grading [%s]", filename)

                grading_result = await self.service.grade(internal)

                logger.info("Done    [%s] — score=%.1f", filename, grading_result.score)

                entry: dict = {
                    "filename": filename,
                    "score": grading_result.score,
                    "feedback": grading_result.feedback,
                }

                if with_reason:
                    entry["reasoning"] = grading_result.reasoning
                    
                return idx, entry
            except Exception as exc:
                logger.error("Error grading [%s]: %s", filename, exc)
                return idx, {"filename": filename, "error": str(exc)}

    @staticmethod
    def _is_supported_zip_entry(name: str) -> bool:
        normalized = name.strip("/")
        if not normalized:
            return False
        if name.endswith("/"):
            return False

        parts = normalized.split("/")
        if "__MACOSX" in parts:
            return False

        return True

    @staticmethod
    def _build_excel(results: list[dict], with_reason: bool) -> bytes:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Grading Results"

        headers = ["No", "Filename", "Score", "Feedback"]
        if with_reason:
            headers.append("Reasoning")
        headers.append("Error")

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, result in enumerate(results, start=2):
            col = 1
            ws.cell(row=row_idx, column=col, value=row_idx - 1)
            col += 1
            ws.cell(row=row_idx, column=col, value=result.get("filename", ""))
            col += 1
            ws.cell(row=row_idx, column=col, value=result.get("score"))
            col += 1
            ws.cell(row=row_idx, column=col, value=result.get("feedback", ""))
            col += 1
            if with_reason:
                ws.cell(row=row_idx, column=col, value=result.get("reasoning"))
                col += 1
            ws.cell(row=row_idx, column=col, value=result.get("error", ""))

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
